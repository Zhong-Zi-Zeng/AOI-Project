from __future__ import annotations
import sys
import os

sys.path.append(os.path.join(os.getcwd()))

import openpyxl
import numpy as np
import pandas as pd
import argparse
from typing import Optional, Tuple, List, Union, Dict
from tqdm import tqdm

import torch
from pycocotools.coco import COCO
from torchvision.ops import nms
from colorama import Fore, Back, Style, init
from rich.console import Console
from rich.table import Table

from engine.general import (get_work_dir_path, save_json, load_json, xywh_to_xyxy)
from engine.timer import TIMER
from engine.builder import Builder
from model_zoo import BaseInstanceModel


def get_args_parser():
    parser = argparse.ArgumentParser('Model evaluation script.', add_help=False)

    parser.add_argument('--config', '-c', type=str, required=True,
                        help='The path of config.')

    parser.add_argument('--excel', '-e', type=str,
                        help='Existing Excel file.'
                             'If given the file, this script will append new value in the given file.'
                             'Otherwise, this script will create a new Excel file depending on the task type.')

    parser.add_argument('--detected_json', '-d', type=str,
                        help='Existing detected file.'
                             'If given the file, this script will be evaluate directly.')

    parser.add_argument('--dir_name', type=str,
                        help='The name of work dir.')

    parser.add_argument('--multi_conf', action="store_true",
                        help='Using multi confidence threshold lie in range 0.3 to 0.9.')

    return parser


class Writer:
    def __init__(self, cfg: dict, excel_path: Optional[str] = None):
        self.cfg = cfg
        self.excel_path = excel_path

        if excel_path is None:
            self.wb = openpyxl.Workbook()
            self.wb.remove(self.wb['Sheet'])
            self.excel_path = os.path.join(get_work_dir_path(cfg), 'result.xlsx')
            self._create_sheet()
            self._append_title()
        else:
            self.wb = openpyxl.load_workbook(self.excel_path)
            self.excel_path = os.path.join(get_work_dir_path(cfg), 'result.xlsx')
            self._append_title()

    def __enter__(self):
        # For common metrics
        self.common_metrics = [self.cfg['model_name'],
                               self.cfg['optimizer'],
                               self.cfg['imgsz'][0],
                               self.cfg['use_patch'],
                               self.cfg['number_of_class'],
                               round(TIMER[2].dt, 3),
                               round(TIMER[3].dt, 3),
                               self.cfg['nms_thres'],
                               self.cfg['conf_thres'],
                               self.cfg['iou_thres'], ]

    def __exit__(self, exc_type, exc_val, exc_tb):
        """自動儲存"""

        self.wb.save(self.excel_path)
        self._save_to_json()

        print('Excel file has been saved to the {}'.format(self.excel_path))

    def _check_has_title(self, sheet_name: str):
        """檢查excel檔是否已經有標題了"""
        work_sheet = self.wb[sheet_name]
        if work_sheet.max_row == 1:
            return False
        return True

    def _save_to_json(self):
        """將excel檔轉換成json後儲存"""

        path = os.path.join(get_work_dir_path(self.cfg), 'result.json')
        excel = pd.read_excel(self.excel_path)
        data_dict = excel.to_dict()
        save_json(path, data_dict, indent=2)

        print('Json file has been saved to the {}'.format(path))

    def _append_title(self):
        """從config中的common、metrics、class_names添加到對應的sheet"""
        for idx, sheet_name in enumerate(self.cfg['sheet_names']):
            if self._check_has_title(sheet_name):
                continue

            if idx == 0:
                self.write_col(self.cfg['common'] + self.cfg['metrics_for_all'], sheet_name, row=1)
            else:
                self.write_col(self.cfg['common'] + self.cfg['class_names'], sheet_name, row=1)

    def _create_sheet(self):
        """依照config中的sheet_names去創建新的work sheet"""
        for idx, sheet_name in enumerate(self.cfg['sheet_names']):
            self.wb.create_sheet(sheet_name)

    def write_col(self,
                  data_list: list[str | float | int],
                  sheet_name: str,
                  row: Optional[int] = None):
        """將資料寫入到excel中"""
        worksheet = self.wb[sheet_name]

        if row is None:
            real_max_row = worksheet.max_row
            while real_max_row > 0:
                if all(cell.value is None for cell in worksheet[real_max_row]):
                    real_max_row -= 1
                else:
                    break
            row = real_max_row + 1

        for col_idx, value in enumerate(data_list, start=1):
            worksheet.cell(row=row, column=col_idx, value=value)


class Evaluator:
    def __init__(self,
                 model: BaseInstanceModel,
                 cfg: dict,
                 detected_json: Optional[str] = None,
                 excel_path: Optional[str] = None):
        self.model = model
        self.cfg = cfg
        self.detected_json = detected_json

        self.coco_gt = COCO(os.path.join(cfg["coco_root"], 'annotations', 'instances_val2017.json'))
        self.writer = Writer(cfg, excel_path=excel_path)

    @classmethod
    def build_by_config(cls, cfg: dict):
        _model = Builder.build_model(cfg)
        return cls(model=_model, cfg=cfg)

    @staticmethod
    def remove_pass_images(coco: COCO):
        """
            Removes image IDs that only contain the 'Pass' class from the COCO dataset.
            Args:
                coco (COCO): COCO dataset object
            Returns:
                all_defect_image_ids (list): List of image IDs that only contain the defect class
                pass_category_id (int): The 'Pass' category ID
                defect_category_ids (list): List of defect category IDs
        """
        # Get all image IDs and category IDs
        all_image_ids = coco.getImgIds()
        all_category_ids = coco.getCatIds()

        # Get all category names
        category_names = [coco.loadCats(cat_id)[0]['name'] for cat_id in all_category_ids]

        # Check if 'Pass' class is present
        assert 'Pass' in category_names, "Pass class not found in COCO dataset"

        # Get the category ID for the 'Pass' class
        pass_category_id = coco.getCatIds(catNms=['Pass'])[0]

        # Get non-pass category IDs
        defect_category_ids = [cat_id for cat_id in all_category_ids if cat_id != pass_category_id]

        # Get image IDs that only contain the 'Pass' class
        pass_images_id = []
        for img_id in coco.getImgIds(catIds=[pass_category_id]):
            annIds = coco.getAnnIds(imgIds=[img_id])
            anns = coco.loadAnns(annIds)
            if all(ann['category_id'] == pass_category_id for ann in anns):
                pass_images_id.append(img_id)

        # Remove image IDs that only contain the 'Pass' class
        all_defect_image_ids = list(set(all_image_ids) - set(pass_images_id))
        all_defect_image_ids.sort()

        return all_defect_image_ids, pass_category_id, defect_category_ids

    def _filter_result(self, score_threshold: float) -> List[dict]:
        """
            這個函數基於指定的分數閾值篩選檢測結果，對篩選後的結果執行非最大值抑制（NMS），並返回篩選和NMS處理後的檢測結果列表。

        Args:
            score_threshold (float): 用於篩選檢測結果的分數閾值。

        Returns:
            result (list[dict]): 經過篩選和非最大值抑制（NMS）處理後的檢測結果列表，包含以下信息：
                - 'image_id'（int）：對應的圖片id。
                - 'category_id'（int）：類別
                - 'bbox'（list）：檢測對象的邊界框坐標，格式為 [x, y, w, h]。
                - 'score'（float）：檢測結果的置信分數，四捨五入到小數點後5位。
        """

        detected_results = load_json(self.detected_json)
        result = []

        # 過濾低於conf_thres的bbox
        filtered_results = [result for result in detected_results if result['score'] >= score_threshold]

        # 對每一張圖片進行nms
        for img_id in self.coco_gt.getImgIds():
            xyxy_bbox_list, xywh_bbox_list, score_list, class_list = [], [], [], []

            # 取出每一張圖片的bbox資訊
            for rs in filtered_results:
                if rs['image_id'] == img_id:
                    xyxy_bbox_list.append(xywh_to_xyxy(rs['bbox'])[0])
                    xywh_bbox_list.append(rs['bbox'])
                    score_list.append(rs['score'])
                    class_list.append(rs['category_id'])

            if class_list and xywh_bbox_list and score_list:
                # 執行nms
                indices = nms(boxes=torch.FloatTensor(np.array(xyxy_bbox_list)),
                              scores=torch.FloatTensor(np.array(score_list)),
                              iou_threshold=self.cfg['nms_thres']).cpu().numpy()

                class_list = np.array(class_list)[indices].tolist()
                xywh_bbox_list = np.array(xywh_bbox_list)[indices].tolist()
                score_list = np.array(score_list)[indices].tolist()

            # 新的檢測結果
            for cls, score, bbox in zip(class_list, score_list, xywh_bbox_list):
                result.append({
                    'image_id': img_id,
                    'category_id': cls,
                    'bbox': bbox,
                    'score': round(score, 5),
                })

        return result

    def _get_iou(self,
                 dt_bbox: Union[list | np.ndarray],
                 gt_bbox: Union[list | np.ndarray]):
        """
            計算dt_bbox與gt_bbox之間的iou

            Args:
                dt_bbox (list | np.ndarray): [D, 4], 格式為[x, y, w, h]
                gt_bbox (list | np.ndarray): [G, 4], 格式為[x, y, w, h]
            Returns:
                Ious np.ndarray(float): [D, G]
        """
        dt_bbox = np.array(dt_bbox, dtype=np.float32)
        gt_bbox = np.array(gt_bbox, dtype=np.float32)

        # xywh -> x1y1x2y2
        dt_bbox = xywh_to_xyxy(dt_bbox)
        gt_bbox = xywh_to_xyxy(gt_bbox)

        lt = np.maximum(dt_bbox[:, None, :2], gt_bbox[:, :2])  # [N,M,2]
        rb = np.minimum(dt_bbox[:, None, 2:], gt_bbox[:, 2:])  # [N,M,2]

        wh = np.maximum(0, rb - lt)  # [N,M,2]
        inter = wh[:, :, 0] * wh[:, :, 1]  # [N,M]

        area_1 = (dt_bbox[:, 2] - dt_bbox[:, 0]) * (dt_bbox[:, 3] - dt_bbox[:, 1])
        area_2 = (gt_bbox[:, 2] - gt_bbox[:, 0]) * (gt_bbox[:, 3] - gt_bbox[:, 1])

        union = area_1[:, None] + area_2 - inter

        iou = inter / union

        return iou

    def _generate_det(self):

        img_id_list = self.coco_gt.getImgIds()
        detected_result = []

        for img_id in tqdm(img_id_list):
            # Load image
            img_info = self.coco_gt.loadImgs([img_id])[0]
            img_file = os.path.join(self.cfg['coco_root'], 'val2017', img_info['file_name'])

            # Inference
            result = self.model.predict(img_file, conf_thres=self.cfg['conf_thres'], nms_thres=self.cfg['nms_thres'])

            class_list = result['class_list']
            score_list = result['score_list']
            bbox_list = result['bbox_list']

            # Analyze result
            if self.cfg['task'] == 'instance_segmentation':
                rle_list = result['rle_list']
                for cls, score, bbox, rle in zip(class_list, score_list, bbox_list, rle_list):
                    detected_result.append({
                        'image_id': img_id,
                        'category_id': cls,
                        'bbox': bbox,
                        'score': round(score, 5),
                        'segmentation': rle
                    })
            # Object detection or Semantic segmentation
            else:
                for cls, score, bbox in zip(class_list, score_list, bbox_list):
                    detected_result.append({
                        'image_id': img_id,
                        'category_id': cls,
                        'bbox': bbox,
                        'score': round(score, 5),
                    })

        # Save detected file
        self.detected_json = os.path.join(get_work_dir_path(self.cfg), 'detected.json')
        save_json(self.detected_json, detected_result, indent=2)

    def eval(self) -> list:
        # Generate detected json
        if self.detected_json is None:
            self._generate_det()

        print('=' * 40)
        print(Fore.BLUE + "Confidence score: {:.1}".format(self.cfg['conf_thres']) + Fore.WHITE)
        print('=' * 40)

        with self.writer:
            # Get all defect image ids
            all_defect_images, pass_category_id, defect_category_ids = self.remove_pass_images(self.coco_gt)

            # Number of defects
            all_category_ids = self.coco_gt.getCatIds()
            all_category_names = [cat['name'] for cat in self.coco_gt.loadCats(all_category_ids)]
            all_defects = sum([len(self.coco_gt.getAnnIds(catIds=[cat_id]))
                               for cat_id in all_category_ids if cat_id != pass_category_id])

            # Filter detection result
            dt_result = self._filter_result(score_threshold=self.cfg["conf_thres"])
            if len(dt_result) == 0:
                print(Fore.RED + 'Can not detect anything! All of the values are zero.' + Fore.WHITE)
                return [0] * 4

            # Load detection result
            coco_dt = self.coco_gt.loadRes(dt_result)

            # Calculate the number of detections and false positives for each image
            nf_recall_img = 0
            nf_fpr_img = 0
            nf_recall_ann = 0
            nf_fpr_ann = 0
            fpr_image_name = []
            each_detect_score = {defect_category_id: [] for defect_category_id in defect_category_ids}

            for img_id in all_defect_images:
                gt_ann = self.coco_gt.loadAnns(self.coco_gt.getAnnIds(imgIds=[img_id]))
                dt_ann = coco_dt.loadAnns(coco_dt.getAnnIds(imgIds=[img_id]))

                # Extract predictions that are classified as defect and ground truth is also defect
                defected_gt_bboxes = np.array([gt['bbox'] for gt in gt_ann if gt['category_id'] != pass_category_id])
                defected_dt_bboxes = np.array([dt['bbox'] for dt in dt_ann if dt['category_id'] != pass_category_id])

                # Check if there is any prediction
                if len(defected_dt_bboxes) == 0 or len(defected_gt_bboxes) == 0:
                    continue

                # Calculate iou
                dt_gt_iou = self._get_iou(defected_dt_bboxes, defected_gt_bboxes)

                # Filter iou
                dt_gt_iou[dt_gt_iou < self.cfg["iou_thres"]] = 0

                # ==========To detected all classes recall and fpr==========
                # Calculate the recall
                nf_recall = np.sum(np.any(dt_gt_iou != 0, axis=1))
                nf_recall = nf_recall if nf_recall < len(defected_gt_bboxes) else len(defected_gt_bboxes)
                nf_recall_ann += nf_recall
                nf_recall_img += 1 if nf_recall > 0 else 0

                # Calculate the fpr
                nf_fpr = np.sum(~np.any(dt_gt_iou != 0, axis=1))
                nf_fpr_ann += nf_fpr
                if nf_fpr > 0:
                    nf_fpr_img += 1
                    fpr_image_name.append(self.coco_gt.loadImgs(ids=[img_id])[0]['file_name'])

                # ==========To detected each class recall and average confidence==========
                gt_class_id = np.array([gt['category_id'] for gt in gt_ann if gt['category_id'] != pass_category_id])
                defected_dt_score = np.array([dt['score'] for dt in dt_ann if dt['category_id'] != pass_category_id])
                class_id = gt_class_id[np.sum(dt_gt_iou, axis=0) != 0]
                dt_score = defected_dt_score[np.any(dt_gt_iou > 0, axis=1)]
                for cls_id, score in zip(class_id, dt_score):
                    each_detect_score[cls_id].append(score)

            result = [
                round((nf_recall_img / len(all_defect_images)) * 100, 2),  # 檢出率 (圖片)
                round((nf_fpr_img / len(all_defect_images)) * 100, 2),  # 過殺率 (圖片)
                nf_recall_ann,  # 檢出數 (瑕疵)
                nf_fpr_ann,  # 過殺數 (瑕疵)
            ]

            # ==========Print information==========
            console = Console()

            # Print result recall and fpr
            table = Table(title="Metrics for all classes", title_justify="left")

            for title, value in zip(self.cfg['metrics_for_all'], result):
                table.add_column(title, justify="center", style="cyan", no_wrap=True)

            table.add_row(*[str(value) for value in result])
            console.print(table)

            # Print each class recall and average confidence
            table = Table(title="Metrics for each class", title_justify="left", show_header=True,
                          header_style="bold magenta")
            table.add_column("Category")
            table.add_column("Total", justify='center', style="cyan", no_wrap=True)
            table.add_column("Recall", justify='center', style="cyan", no_wrap=True)
            table.add_column("Avg confidence", justify='center', style="cyan", no_wrap=True)
            table.add_column("Recall rate", justify='center', style="cyan", no_wrap=True)

            for cls_id, scores in each_detect_score.items():
                cat_name = all_category_names[cls_id]
                total = len(self.coco_gt.getAnnIds(catIds=[cls_id]))
                avg_confidence = f"{np.mean(scores):.3f}" if scores else "-"
                recall = len(scores)
                recall_rate = f"{recall / total:.3f}" if total > 0 else "-"

                table.add_row(
                    cat_name, str(total), str(recall), avg_confidence, recall_rate
                )

            console.print(table)

            # Print information
            table = Table(title="Result Analysis", show_lines=True, title_justify="left")
            table.add_column("Information")
            table.add_column("Value")

            table.add_row("Number of defect image", str(len(all_defect_images)))
            table.add_row("Number of defect", str(all_defects))
            table.add_row("FPR images", ", ".join(fpr_image_name))

            console.print(table)

            # Print timer
            print('\n\n')
            for timer in TIMER:
                print(f"{timer.name:15s} {timer.dt:.3f}", end=' | ')
            print('\n\n')

            # Store value
            self.writer.write_col(self.writer.common_metrics +
                                  result,
                                  sheet_name=self.cfg['sheet_names'][0])

            return result


if __name__ == "__main__":
    parser = argparse.ArgumentParser('Model evaluation script.',
                                     parents=[get_args_parser()])
    args = parser.parse_args()

    # Create builder
    builder = Builder(config_path=args.config, task='eval', work_dir_name=args.dir_name)

    # Build config
    cfg = builder.build_config()

    # Build model
    model = builder.build_model(cfg)

    # Use multi confidence threshold
    detected_json = args.detected_json
    if args.multi_conf:
        confidences = np.arange(0.3, 1.0, 0.1)
    else:
        confidences = [cfg['conf_thres']]

    results_df = pd.DataFrame(
        columns=['Confidence Threshold', 'Recall (Image)', 'FPR (Image)', 'Recall (Defect)', 'FPR (Defect)'])

    for idx, conf in enumerate(confidences):
        cfg['conf_thres'] = conf

        # Build evaluator
        evaluator = Evaluator(model=model, cfg=cfg, excel_path=args.excel, detected_json=detected_json)
        _result = evaluator.eval()

        results_df.loc[idx] = [conf, _result[0], _result[1], _result[2], _result[3]]

        # Save detected result
        if idx == 0:
            detected_json = evaluator.detected_json

    results_df['Recall (Image)'] = results_df['Recall (Image)'].apply(lambda x: f"{x}%")
    results_df['FPR (Image)'] = results_df['FPR (Image)'].apply(lambda x: f"{x}%")
    transposed_results_df = results_df.set_index('Confidence Threshold').T
    transposed_results_df.to_excel('evaluation_results_transposed.xlsx')
